#Exercise: show the elements of a document excel as a dictionary

#install the package openpyxl for read/write document excel
#pip install openpyxl

import openpyxl 

inv_file = openpyxl.load_workbook("inventory.xlsx") #load the file of excel
product_list =  inv_file["inventory"] #put the name of the sheet of excel

#show the element of the document as a dictionary
products_per_supplier = {}
    
#use looping for go through each and every row in the sheet and execute same logic 
#on each item and repeat the loop the number of row of the table of elements

#max_row show the number maximum of row for element of the sheet
#range - create a sequence of integer numbers (value valid for loop)
for product_row in range(2, product_list.max_row + 1): #is two because the list of element start in the row two and plus one 
    #is for the loop only count until product_row < product_list.max_row and we must include the last element.
    #now we can see the element of each cell  
    element_cell = product_list.cell(product_row, 4) #the second parameter is the number of columns of the document of excel
    #ncrease between each row plus one
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_num_products + 1

    else:
        products_per_supplier[supplier_name] = 1

#show the list of each company
print(product_list) #print elements of the excel as a dictionary
